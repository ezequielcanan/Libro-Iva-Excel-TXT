import openpyxl as xl
from datetime import datetime

excelFilePath = "./IVA VENTAS - Guzman 10-2023.xlsx"

workbookObject = xl.load_workbook(excelFilePath)
sheetObject = workbookObject.active

minRow = sheetObject.min_row
maxRow = sheetObject.max_row

txtFile = open("LIBRO_IVA_DIGITAL_VENTAS_CBTE.txt", "x", encoding="cp1252")
alicuotasFile = open("ALICUOTAS.txt", "x", encoding="cp1252")

tiposComprobantes = {
  "TCK B": "006",
  "FAC A": "001",
  "TCK  ": "100"
}


def writeData(row):
  for col in row:
    colCoordinate = col.coordinate[0]

    if col.value is None and colCoordinate == "A":
      break
    
    if row[1].value == "TCK  ":
      break

    match colCoordinate:
      case "A":
        date = datetime.strptime(col.value, "%d/%m/%Y").date().strftime("%Y%m%d")
        txtFile.write(date)
      case "B":
        tipoComprobante = tiposComprobantes[col.value]
        txtFile.write(f"{tipoComprobante}00001")
        alicuotasFile.write(f"{tipoComprobante}00001")
      case "C":
        nroComprobante = col.value.replace("-","").strip().zfill(20)
        txtFile.write(f"{nroComprobante}{nroComprobante}")
        alicuotasFile.write(f"{nroComprobante}{str(row[6].value).replace('.','').zfill(15)}0005{str(round(row[5].value-row[6].value,2)).replace('.','').zfill(15)}\n")
      case "D":
        if col.value == "CONSUMIDOR FINAL                                  ":
          txtFile.write(f"99{'0'.zfill(20)}{'VENTA GLOBAL DIARIA'.ljust(30)[0:30]}")
      case "E":
        if col.value is not None:
          if all(c == col.value[0] for c in col.value):
            txtFile.write(f"99{'0'.zfill(20)}{row[3].value.ljust(30)[0:30]}")
          else:
            txtFile.write(f"80{col.value.replace('-','').strip().zfill(20)}{row[3].value.ljust(30)[0:30]}")
      case "F":
        total = str(col.value).zfill(13)+str(col.value-int(col.value))
        txtFile.write(f"{total}{'0'.zfill(106)}PES00010000001 000000000000000{datetime.strptime(row[0].value, '%d/%m/%Y').date().strftime('%Y%m%d')}\n")



for row in sheetObject.iter_rows(min_row = 2):
  writeData(row)

txtFile.close()
alicuotasFile.close()